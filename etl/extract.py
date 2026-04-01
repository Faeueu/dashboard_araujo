"""
extract.py — Leitura das planilhas VR Master e extração em dicts limpos.
Cada função retorna uma lista de dicts prontos para transformação.
"""
import os
import re
from openpyxl import load_workbook
from config import (
    DADOS_DIR, FONTES, MESES_ARQUIVOS, PASTA_LOJA,
    COLS_VENDAS, COLS_DRO, COLS_ESTOQUE, COLS_CADASTRO,
    COLS_ENTRADA_SAIDA, COLS_INVENTARIO, COLS_VALIDADE,
)


def _read_sheet(filepath, col_map=None, skip_empty_key=None):
    """
    Lê uma planilha xlsx e retorna lista de dicts.
    - col_map: {coluna_excel: nome_interno} para renomear
    - skip_empty_key: pula linhas onde essa coluna interna está vazia
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Primeira linha = cabeçalhos
    headers = [str(h).strip() if h else f'_col_{i}' for i, h in enumerate(rows[0])]
    data = []

    for row in rows[1:]:
        raw = {}
        for i, val in enumerate(row):
            if i < len(headers):
                raw[headers[i]] = val
        data.append(raw)

    if col_map:
        mapped = []
        for row in data:
            m = {}
            for excel_col, internal_name in col_map.items():
                # Pega primeiro match (cuidado com "Descrição" duplicado)
                if excel_col in row:
                    m[internal_name] = row[excel_col]
                else:
                    m[internal_name] = None
            mapped.append(m)
        data = mapped

    if skip_empty_key:
        data = [d for d in data if d.get(skip_empty_key)]

    return data


def _parse_pct(val):
    """Converte '20,90 % ' → 20.90 como float"""
    if val is None:
        return 0.0
    s = str(val).strip().replace('%', '').replace(',', '.').strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_num(val):
    """Converte string numérica com possíveis vírgulas para float"""
    if val is None:
        return 0.0
    s = str(val).strip()
    if not s or s == '-':
        return 0.0
    # Se tem vírgula como decimal (padrão BR)
    if ',' in s and '.' not in s:
        s = s.replace(',', '.')
    elif ',' in s and '.' in s:
        # 1.234,56 → 1234.56
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date_vr(val):
    """Converte '01/04/2026 - QUA' → '2026-04-01'"""
    if not val:
        return None
    s = str(val).strip()
    match = re.match(r'(\d{2})/(\d{2})/(\d{4})', s)
    if match:
        d, m, y = match.groups()
        return f"{y}-{m}-{d}"
    return None


# ════════════════════════════════════════════════════════════════
#  EXTRATORES POR FONTE
# ════════════════════════════════════════════════════════════════

def extrair_vendas(loja_pasta: str, mes: str) -> list[dict]:
    """
    Extrai vendas do período para uma loja/mês.
    Retorna lista de dicts com campos normalizados.
    """
    arquivo = MESES_ARQUIVOS.get(mes)
    if not arquivo:
        print(f"  ⚠ Mês {mes} não mapeado")
        return []

    path = os.path.join(DADOS_DIR, FONTES['vendas'], loja_pasta, arquivo)
    if not os.path.exists(path):
        print(f"  ⚠ Arquivo não encontrado: {path}")
        return []

    print(f"  📄 Lendo vendas: {path}")
    rows = _read_sheet(path)

    result = []
    for row in rows:
        codigo = str(row.get('Produto', '')).strip()
        if not codigo or not re.match(r'^\d+', codigo):
            continue

        result.append({
            'codigo':            codigo,
            'descricao':         str(row.get('Descrição', '')).strip(),
            'cod_barras':        str(row.get('Código de Barras', '')).strip(),
            'quantidade':        _parse_num(row.get('Quantidade')),
            'venda':             _parse_num(row.get('Venda')),
            'custo_com_imposto': _parse_num(row.get('Custo c/ Imposto')),
            'custo_sem_imposto': _parse_num(row.get('Custo s/ Imposto')),
            'margem_bruta_pct':  _parse_pct(row.get('Margem Bruta')),
            'lucro':             _parse_num(row.get('Lucro')),
            'icms':              _parse_num(row.get('ICMS')),
            'pis_cofins':        _parse_num(row.get('PIS/COFINS')),
            'dias_efetivos':     int(_parse_num(row.get('Dias Efetivos Venda', 0))),
            'loja_pasta':        loja_pasta,
            'mes':               mes,
        })

    print(f"    → {len(result)} produtos extraídos")
    return result


def extrair_dro(loja_pasta: str, mes: str) -> list[dict]:
    """
    Extrai DRO (Demonstrativo). Formato especial: 2 linhas de cabeçalho.
    """
    arquivo = MESES_ARQUIVOS.get(mes)
    if not arquivo:
        return []

    # DRO usa nomes diferentes para março
    nome_arquivo = arquivo.replace('março', 'marco')
    path = os.path.join(DADOS_DIR, FONTES['dro'], loja_pasta, nome_arquivo)
    if not os.path.exists(path):
        # Tenta o nome original
        path = os.path.join(DADOS_DIR, FONTES['dro'], loja_pasta, arquivo)
    if not os.path.exists(path):
        print(f"  ⚠ DRO não encontrado para {loja_pasta}/{mes}")
        return []

    print(f"  📄 Lendo DRO: {path}")
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # DRO: linha 0 = "", "Realizado", "", "Previsto", ""
    #       linha 1 = "Descrição", "Valor", "Percentual", "Valor", "Percentual"
    #       linhas 2+ = dados
    result = []
    for row in rows[2:]:  # Skip 2 header rows
        descricao = str(row[0] or '').strip()
        if not descricao:
            continue
        result.append({
            'descricao':       descricao,
            'valor_realizado': _parse_num(row[1]),
            'percentual':      _parse_pct(row[2]),
            'loja_pasta':      loja_pasta,
            'mes':             mes,
        })

    print(f"    → {len(result)} linhas DRO extraídas")
    return result


def extrair_cadastro(loja_pasta: str) -> list[dict]:
    """
    Extrai cadastro de produtos. Não é mensal — um snapshot.
    """
    # Procura arquivo na pasta
    pasta = os.path.join(DADOS_DIR, FONTES['cadastro'], loja_pasta)
    if not os.path.exists(pasta):
        print(f"  ⚠ Pasta cadastro não encontrada: {pasta}")
        return []

    arquivos = [f for f in os.listdir(pasta) if f.endswith('.xlsx')]
    if not arquivos:
        return []

    path = os.path.join(pasta, arquivos[0])
    print(f"  📄 Lendo cadastro: {path}")
    rows = _read_sheet(path)

    result = []
    # Cuidado: "Descrição" aparece 2x no cabeçalho (col 2 e col 20)
    # A primeira é a descrição do produto, a segunda do NCM/Nat. Receita
    for row in rows:
        codigo = str(row.get('Código', '')).strip()
        if not codigo or codigo == 'NÃO':
            continue

        result.append({
            'codigo':        codigo,
            'descricao':     str(row.get('Descrição', '')).strip(),
            'embalagem':     str(row.get('Embalagem', '')).strip(),
            'ncm':           str(row.get('NCM', '')).strip(),
            'mercadologico': str(row.get('Mercadológico', '')).strip(),
            'situacao':      str(row.get('Situação', 'ATIVO')).strip(),
            'familia':       str(row.get('Família', '')).strip(),
        })

    print(f"    → {len(result)} produtos cadastrados")
    return result


def extrair_estoque(loja_pasta: str) -> list[dict]:
    """
    Extrai posição de estoque. Formato especial: linhas alternadas
    (produto + sublinha de loja).
    """
    pasta = os.path.join(DADOS_DIR, FONTES['estoque'], loja_pasta)
    if not os.path.exists(pasta):
        print(f"  ⚠ Pasta estoque não encontrada: {pasta}")
        return []

    arquivos = [f for f in os.listdir(pasta) if f.endswith('.xlsx')]
    if not arquivos:
        return []

    path = os.path.join(pasta, arquivos[0])
    print(f"  📄 Lendo estoque: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = [str(h).strip() if h else '' for h in rows[0]]
    result = []

    # Formato: linha ímpar = produto (código + descricao na col Produto)
    #          linha par   = loja (indentada na col Produto, com valores)
    current_produto = None
    for row in rows[1:]:
        col0 = str(row[0] or '').strip()

        if not col0:
            continue

        # Linha de produto: tem código e descrição juntos (ex: "917712 -  CALDO...")
        if re.match(r'^\d+\s*-\s*', col0):
            match = re.match(r'^(\d+)\s*-\s*(.*)', col0)
            if match:
                current_produto = {
                    'codigo': match.group(1).strip(),
                    'descricao': match.group(2).strip(),
                    'embalagem': str(row[1] or '').strip(),
                    'cod_barras': str(row[2] or '').strip(),
                    'quantidade': _parse_num(row[15]) if len(row) > 15 else 0,
                }
            continue

        # Linha de loja: indentada, contém valores
        loja_nome = col0.strip().upper()
        if current_produto and loja_nome in ('PRIMAVERA L01', 'ARAUJO L02', 'MARAUTO L03'):
            result.append({
                'codigo':            current_produto['codigo'],
                'descricao':         current_produto['descricao'],
                'embalagem':         current_produto['embalagem'],
                'cod_barras':        current_produto['cod_barras'],
                'loja_raw':          loja_nome,
                'preco_venda':       _parse_num(row[3]),
                'total_venda':       _parse_num(row[4]),
                'custo_com_imposto': _parse_num(row[5]),
                'custo_sem_imposto': _parse_num(row[6]),
                'curva':             str(row[9] or '-').strip(),
                'quantidade':        _parse_num(row[15]) if len(row) > 15 else 0,
            })

    print(f"    → {len(result)} linhas estoque extraídas")
    return result


def extrair_entrada_saida(loja_pasta: str, mes: str) -> list[dict]:
    """Extrai movimentação entrada/saída para uma loja/mês."""
    arquivo = MESES_ARQUIVOS.get(mes)
    if not arquivo:
        return []

    path = os.path.join(DADOS_DIR, FONTES['entrada_saida'], loja_pasta, arquivo)
    if not os.path.exists(path):
        print(f"  ⚠ Entrada/Saída não encontrado: {path}")
        return []

    print(f"  📄 Lendo entrada/saída: {path}")
    rows = _read_sheet(path)

    result = []
    for row in rows:
        codigo = str(row.get('Produto', '')).strip()
        if not codigo or not re.match(r'^\d+', codigo):
            continue

        result.append({
            'codigo':        codigo,
            'descricao':     str(row.get('Descrição', '')).strip(),
            'qtd_entrada':   _parse_num(row.get('Qtd. Entrada')),
            'valor_entrada': _parse_num(row.get('Entrada')),
            'qtd_saida':     _parse_num(row.get('Qtd. Saída')),
            'valor_saida':   _parse_num(row.get('Saída')),
            'qtd_venda':     _parse_num(row.get('Qtd. Venda')),
            'valor_venda':   _parse_num(row.get('Venda')),
            'custo_venda':   _parse_num(row.get('Custo Venda')),
            'estoque_atual': _parse_num(row.get('Estoque Atual')),
            'loja_pasta':    loja_pasta,
            'mes':           mes,
        })

    print(f"    → {len(result)} movimentações extraídas")
    return result


def extrair_inventario(loja_pasta: str, mes: str) -> list[dict]:
    """Extrai inventário mensal."""
    arquivo = MESES_ARQUIVOS.get(mes)
    if not arquivo:
        return []

    path = os.path.join(DADOS_DIR, FONTES['inventario'], loja_pasta, arquivo)
    if not os.path.exists(path):
        print(f"  ⚠ Inventário não encontrado: {path}")
        return []

    print(f"  📄 Lendo inventário: {path}")
    rows = _read_sheet(path)

    result = []
    for row in rows:
        codigo = str(row.get('Código', '')).strip()
        sel = str(row.get('Selecionado', '')).strip()
        if sel not in ('SIM', 'NÃO') or not codigo:
            continue

        result.append({
            'codigo':            codigo,
            'descricao':         str(row.get('Descrição', '')).strip(),
            'data_ref':          str(row.get('Data', '')).strip(),
            'embalagem':         str(row.get('Embalagem', '')).strip(),
            'quantidade':        _parse_num(row.get('Quantidade')),
            'preco_venda':       _parse_num(row.get('Preço Venda')),
            'custo_com_imposto': _parse_num(row.get('Custo c/ Imposto')),
            'custo_sem_imposto': _parse_num(row.get('Custo s/ Imposto')),
            'loja_raw':          str(row.get('Loja', '')).strip(),
            'loja_pasta':        loja_pasta,
            'mes':               mes,
        })

    print(f"    → {len(result)} itens inventariados")
    return result


def extrair_validade() -> list[dict]:
    """
    Extrai controle de validade — arquivo único com todas as lojas.
    """
    pasta = os.path.join(DADOS_DIR, FONTES['validade'])
    arquivos = [f for f in os.listdir(pasta) if f.endswith('.xlsx')]
    if not arquivos:
        print("  ⚠ Nenhum arquivo de validade encontrado")
        return []

    path = os.path.join(pasta, arquivos[0])
    print(f"  📄 Lendo validade: {path}")
    rows = _read_sheet(path)

    result = []
    for row in rows:
        codigo = str(row.get('Produto', '')).strip()
        if not codigo or not re.match(r'^\d+', codigo):
            continue

        result.append({
            'codigo':          codigo,
            'descricao':       str(row.get('Descrição', '')).strip(),
            'data_validade':   _parse_date_vr(row.get('Data Validade')),
            'data_entrada':    _parse_date_vr(row.get('Data Entrada')),
            'quantidade':      _parse_num(row.get('Quantidade')),
            'estoque_atual':   _parse_num(row.get('Estoque Atual')),
            'media_venda':     _parse_num(row.get('Média')),
            'ddv':             int(_parse_num(row.get('DDV', 0))),
            'dias_restantes':  int(_parse_num(row.get('Dias Restantes', 0))),
            'risco_vencer':    _parse_pct(row.get('Estoque em Alerta')),
            'custo_observacao': _parse_num(row.get('Custo em Observação')),
            'venda_risco':     _parse_num(row.get('Venda em Risco')),
            'situacao':        str(row.get('Situação', '')).strip(),
            'loja_raw':        str(row.get('Loja', '')).strip(),
        })

    print(f"    → {len(result)} registros de validade extraídos")
    return result


# ════════════════════════════════════════════════════════════════
#  TESTE RÁPIDO
# ════════════════════════════════════════════════════════════════
if __name__ == '__main__':
    print("=" * 60)
    print("TESTE DE EXTRAÇÃO — Primavera Jan/2026")
    print("=" * 60)

    vendas = extrair_vendas('primavera', '2026-01')
    print(f"\nVendas: {len(vendas)} linhas")
    if vendas:
        print(f"  Ex: {vendas[0]}")

    dro = extrair_dro('primavera', '2026-01')
    print(f"\nDRO: {len(dro)} linhas")
    if dro:
        print(f"  Ex: {dro[0]}")

    cadastro = extrair_cadastro('primavera')
    print(f"\nCadastro: {len(cadastro)} linhas")
    if cadastro:
        print(f"  Ex: {cadastro[0]}")

    estoque = extrair_estoque('primavera')
    print(f"\nEstoque: {len(estoque)} linhas")
    if estoque:
        print(f"  Ex: {estoque[0]}")

    es = extrair_entrada_saida('primavera', '2026-01')
    print(f"\nEntrada/Saída: {len(es)} linhas")

    inv = extrair_inventario('primavera', '2026-01')
    print(f"\nInventário: {len(inv)} linhas")

    val = extrair_validade()
    print(f"\nValidade: {len(val)} linhas")

    print("\n✅ Extração concluída com sucesso!")
